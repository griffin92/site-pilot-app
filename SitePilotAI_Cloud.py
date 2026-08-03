import streamlit as st
from google import genai
from google.genai import types
from PIL import Image
from pdf2image import convert_from_bytes, pdfinfo_from_bytes
import openpyxl
import json
import os
import csv
import io
import gc
from datetime import datetime, date, timedelta
from fpdf import FPDF 
from fpdf.enums import XPos, YPos

# ==========================================
# 1. PAGE CONFIG (MUST BE FIRST COMMAND)
# ==========================================
st.set_page_config(page_title="Site Pilot AI", layout="wide", page_icon="🏗️")

# ==========================================
# 2. SETUP & SECURE API CONFIGURATION
# ==========================================
try:
    ai_client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])
except Exception:
    st.error("🚨 CONFIGURATION ERROR: GEMINI_API_KEY not found in Streamlit Cloud Secrets. Please add it to your app settings.")
    st.stop()

# Single source of truth for the model name. gemini-2.5-pro is being retired
# (Oct 2026) and was slower for this workload anyway. gemini-3.6-flash is
# Google's current fast/workhorse tier -- built for latency and multi-step
# jobs like ours, and cheaper per token than 2.5-pro was.
MODEL_NAME = "gemini-3.6-flash"

# ==========================================
# 3. SITE PILOT DESIGN SYSTEM
# ==========================================
# Palette, type, and layout language pulled from jobsite/blueprint vernacular
# rather than generic SaaS defaults: structural navy, steel gray-blue, safety
# orange, and a title-block signature element modeled on the info box found
# on real architectural/construction drawings.
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=IBM+Plex+Sans:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

    :root {
        --sp-navy: #0E1B2E;
        --sp-navy-2: #1B2A42;
        --sp-steel: #52627A;
        --sp-ink: #10151C;
        --sp-safety: #E8590C;
        --sp-verified: #1F7A4D;
        --sp-border: rgba(16, 27, 46, 0.14);
    }

    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
    h1, h2, h3 { font-family: 'Space Grotesk', sans-serif; letter-spacing: -0.02em; }

    /* ---- Buttons: stamped, technical, not rounded-pill default ---- */
    .stButton>button {
        border-radius: 3px;
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        font-size: 0.82em;
        letter-spacing: 0.04em;
        text-transform: uppercase;
        background-color: var(--sp-safety);
        color: #ffffff;
        border: none;
        width: 100%;
        height: 2.9em;
        transition: filter 0.15s ease, box-shadow 0.15s ease;
        box-shadow: none;
    }
    .stButton>button:hover { filter: brightness(1.08); transform: none; box-shadow: 0 2px 10px rgba(232, 89, 12, 0.28); }
    .btn-clear>button { background-color: transparent; color: var(--sp-steel); border: 1px solid var(--sp-border); }
    .btn-clear>button:hover { background-color: rgba(82, 98, 122, 0.08); filter: none; box-shadow: none; }

    /* ---- Sidebar: dark nav rail ---- */
    section[data-testid="stSidebar"] { background-color: var(--sp-navy); border-right: 1px solid rgba(255,255,255,0.06); }
    /* Light text for labels/headings/captions sitting directly on the navy
       background -- deliberately NOT a blanket "*" rule, since that also
       forced light text onto native controls (like the file uploader's
       Browse button) that keep their own light background, killing contrast. */
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] small,
    section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] { color: #E8ECF1; }
    /* Native buttons/controls (file uploader Browse button, etc.) keep
       their own light background -- force dark text so it stays legible
       instead of inheriting the light-on-light from the rule above. */
    section[data-testid="stSidebar"] button:not(.stButton button):not([kind="primary"]) { color: var(--sp-ink) !important; }
    section[data-testid="stSidebar"] button:not(.stButton button):not([kind="primary"]) * { color: var(--sp-ink) !important; }
    section[data-testid="stSidebar"] h2, section[data-testid="stSidebar"] h3 { font-family: 'Space Grotesk', sans-serif; }
    section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1); }

    /* ---- Tabs: numbered workflow phases ---- */
    div[data-testid="stTabs"] button[data-baseweb="tab"] {
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        font-size: 0.8em;
        letter-spacing: 0.03em;
        text-transform: uppercase;
        color: var(--sp-steel);
    }
    div[data-testid="stTabs"] button[aria-selected="true"] { color: var(--sp-ink); border-bottom-color: var(--sp-safety) !important; }

    /* ---- Tool cards: spec-sheet panels ---- */
    .tool-card { padding: 22px 25px; border-radius: 4px; border: 1px solid var(--sp-border); border-top: 3px solid var(--sp-steel); background-color: var(--secondary-background-color); box-shadow: 0 1px 2px rgba(16,27,46,0.05); margin-bottom: 20px; color: var(--text-color); }
    .module-tag { font-family: 'IBM Plex Mono', monospace; font-size: 0.68em; letter-spacing: 0.08em; text-transform: uppercase; color: var(--sp-steel); margin-bottom: 10px; padding-bottom: 8px; border-bottom: 1px dashed var(--sp-border); }

    /* ---- Report output ---- */
    .report-box { padding: 18px 20px; border-radius: 3px; background-color: rgba(82, 98, 122, 0.07); border-left: 3px solid var(--sp-steel); color: var(--text-color); margin-top: 14px; font-size: 0.94em; overflow-x: auto; }
    .report-box.accent-safety { border-left-color: var(--sp-safety); }
    .report-box.accent-verified { border-left-color: var(--sp-verified); }
    .ref-header { background-color: var(--sp-navy); color: #ffffff !important; padding: 7px 14px; border-radius: 3px 3px 0 0; font-weight: 600; font-family: 'IBM Plex Mono', monospace; font-size: 0.75em; letter-spacing: 0.06em; text-transform: uppercase; }

    /* ---- Title block: signature element, modeled on a drawing title block ---- */
    .title-block { display: flex; border: 1.5px solid var(--sp-ink); border-radius: 2px; overflow: hidden; margin-bottom: 26px; background-color: var(--secondary-background-color); flex-wrap: wrap; }
    .tb-field { padding: 10px 18px; border-right: 1px solid var(--sp-border); display: flex; flex-direction: column; justify-content: center; min-width: 140px; }
    .tb-field:last-child { border-right: none; flex: 1; }
    .tb-label { font-family: 'IBM Plex Mono', monospace; font-size: 0.62em; letter-spacing: 0.08em; text-transform: uppercase; color: var(--sp-steel); margin-bottom: 2px; }
    .tb-value { font-family: 'IBM Plex Mono', monospace; font-size: 0.92em; font-weight: 600; color: var(--sp-ink); }
    .tb-brand { background-color: var(--sp-navy); justify-content: center; align-items: center; min-width: 66px; border-right: 1.5px solid var(--sp-ink); }
    .tb-brand .tb-value { color: #ffffff; font-family: 'Space Grotesk', sans-serif; font-size: 1.05em; }

    /* ---- Hero / landing ---- */
    .hero-eyebrow { font-family: 'IBM Plex Mono', monospace; font-size: 0.78em; letter-spacing: 0.14em; text-transform: uppercase; color: var(--sp-safety); margin-bottom: 14px; }
    .hero-title { font-family: 'Space Grotesk', sans-serif; font-size: 3.1em; font-weight: 700; color: var(--sp-ink); margin-bottom: 6px; letter-spacing: -0.03em; }
    .hero-sub { font-size: 1.12em; opacity: 0.75; color: var(--text-color); margin-bottom: 40px; font-weight: 400; max-width: 520px; margin-left: auto; margin-right: auto; }

    .section-title { font-size: 1.22em; font-weight: 600; color: var(--text-color); margin-bottom: 5px; border-bottom: 2px solid var(--sp-border); padding-bottom: 8px; letter-spacing: -0.01em; }
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 4. CLOUD-READY UTILITIES
# ==========================================
@st.cache_resource
def get_pdf_info(file_bytes):
    return pdfinfo_from_bytes(file_bytes)["Pages"]

@st.cache_data
def convert_single_page(file_bytes, page_num):
    # Ram protection: limits to 1600px width
    return convert_from_bytes(file_bytes, first_page=page_num, last_page=page_num, size=(1600, None))[0]

# ==========================================
# GANTT CHART GENERATION (Vertex42 template fill)
# ==========================================
# Fills the boss's exact Vertex42 Gantt template rather than generating a
# generic spreadsheet -- every formula in the template (End date, Cal Days,
# Days Done/Left) stays exactly as-is. We only write the plain input cells:
# Project Start Date (G6) and, per task row, WBS/Task/Predecessors/Start/
# Work Days/% Done. Start dates are computed here in Python (not by the AI)
# since LLMs are unreliable at business-day arithmetic.
GANTT_TEMPLATE_PATH = "templates/gantt_template.xlsx"
GANTT_TASK_START_ROW = 14
GANTT_MAX_TASK_ROWS = 50  # template extended to support WBS rows 14-63

def add_business_days(start_date, work_days):
    """Mirrors the template's WORKDAY(start, work_days-1) formula."""
    if work_days <= 1:
        return start_date
    d = start_date
    remaining = work_days - 1
    while remaining > 0:
        d += timedelta(days=1)
        if d.weekday() < 5:
            remaining -= 1
    return d

def next_business_day(d):
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def schedule_tasks(project_start, tasks):
    """Computes start/end dates sequentially from predecessors."""
    end_by_wbs = {}
    for t in tasks:
        preds = t.get("predecessors") or []
        pred_ends = [end_by_wbs[p] for p in preds if p in end_by_wbs]
        start = next_business_day(max(pred_ends) + timedelta(days=1)) if pred_ends else project_start
        end = add_business_days(start, max(1, t.get("work_days", 1)))
        t["start"] = start
        t["end"] = end
        end_by_wbs[t["wbs"]] = end
    return tasks

def extract_tasks_json(schedule_text):
    """Converts the free-text AI timeline into a structured task list using
    JSON mode -- far more reliable than parsing CSV text the model wrote."""
    sys_prompt = "You are a data extraction engine. Convert construction schedules into structured task lists."
    usr_prompt = f"""Extract every task from this construction timeline into a JSON array.
Each item must have exactly these fields:
- "wbs": sequential integer starting at 1
- "task": short task name (string)
- "work_days": estimated duration in working days (integer, minimum 1)
- "predecessors": array of WBS integers this task depends on (empty array if none)

Output ONLY the raw JSON array, no markdown fences, no other text.

Timeline:
{schedule_text}"""
    response = ai_client.models.generate_content(
        model=MODEL_NAME,
        contents=[usr_prompt],
        config=types.GenerateContentConfig(
            system_instruction=sys_prompt,
            temperature=0.1,
            response_mime_type="application/json"
        )
    )
    raw = response.text.strip().replace('```json', '').replace('```', '').strip()
    return json.loads(raw)

def fill_gantt_template(project_start, tasks):
    """Opens the bundled Vertex42 template and writes only the input cells,
    leaving every locked formula untouched. Returns the .xlsx as bytes."""
    wb = openpyxl.load_workbook(GANTT_TEMPLATE_PATH)
    ws = wb["GanttChart"]
    ws["G6"] = project_start

    scheduled = schedule_tasks(project_start, tasks)
    truncated = len(scheduled) > GANTT_MAX_TASK_ROWS
    if truncated:
        scheduled = scheduled[:GANTT_MAX_TASK_ROWS]

    for i in range(GANTT_MAX_TASK_ROWS):
        row = GANTT_TASK_START_ROW + i
        if i < len(scheduled):
            t = scheduled[i]
            ws.cell(row=row, column=1).value = t["wbs"]
            ws.cell(row=row, column=2).value = t["task"]
            ws.cell(row=row, column=4).value = ",".join(str(p) for p in t.get("predecessors", [])) or None
            ws.cell(row=row, column=7).value = t["start"]
            ws.cell(row=row, column=9).value = max(1, t.get("work_days", 1))
            ws.cell(row=row, column=10).value = 0
        else:
            for col in (1, 2, 4, 7, 9, 10):
                ws.cell(row=row, column=col).value = None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), truncated

# ==========================================
# DRAWING Q&A ENGINE
# ==========================================
# Targeted lookup, as opposed to the one-shot deep-scan engines above.
# Two-stage design:
#   1. (Optional) Route: a cheap TEXT-ONLY call reads the sheet index and
#      picks which sheets likely hold the answer. Skipped if the Auto-Index
#      hasn't been run, since generic "Page 7" titles carry no signal.
#   2. Answer: only the routed sheets get converted to images and sent.
# This keeps a 100+ sheet set from being brute-forced on every question --
# faster, cheaper, and more accurate, since the model isn't wading through
# 90 irrelevant sheets looking for a walk-in cooler.

QA_SYSTEM_PROMPT = """You are a Veteran Commercial Construction Superintendent and Project Manager with 25+ years in the field, reading a set of construction drawings to answer a specific question from the field team.

HOW YOU READ DRAWINGS:
- You read floor plans, RCPs, elevations, sections, details, schedules (door/window/finish/equipment/panel), keynotes, general notes, and legends.
- You cross-reference: a room's finish comes from the Finish Schedule keyed to the room tag; equipment power comes from the Equipment Schedule cross-referenced to the Panel Schedule; dimensions come from dimension strings, not from eyeballing.
- You know equipment tags (WIC = walk-in cooler, RTU, MAU, EF, WH, etc.), and you connect a tag on a plan to its row in the corresponding schedule.

ACCURACY RULES -- THESE ARE ABSOLUTE:
1. NEVER invent, estimate, or infer a number that is not on the drawings. A wrong dimension causes real money and rework in the field.
2. Distinguish clearly between:
   - LABELED: the value is explicitly printed on the drawing (dimension string, schedule cell, note). State it plainly.
   - DERIVED: you calculated it from labeled values (e.g. area = labeled length x labeled width). Show the math and label it as calculated.
   - NOT SHOWN: the information is not on the sheets provided. Say so directly.
3. NEVER scale off a drawing to produce a dimension. If a dimension isn't labeled, say it isn't labeled and name which sheet would likely carry it.
4. ALWAYS cite the sheet you got each fact from, using the sheet number/name given in the sheet label.
5. If sheets conflict with each other, say so explicitly and flag it as a coordination issue -- do not silently pick one.
6. If the question can't be answered from the sheets provided, say exactly that and name which sheet type would have it (e.g. "Not on these sheets -- check the Equipment Schedule, typically on M-601 or E-601").

TONE: Direct, field-practical, brief. Answer like a superintendent briefing another superintendent. Lead with the answer, then the supporting reference. No preamble, no restating the question."""

def select_relevant_sheets(question, drawing_index, candidate_pages, max_sheets=8):
    """Cheap text-only routing pass: picks likely relevant sheets by title.
    Returns None if the index has no real titles (Auto-Index not run yet)."""
    named = [p for p in candidate_pages
             if not str(drawing_index.get(str(p), "")).strip().lower().startswith("page ")]
    if len(named) < max(3, len(candidate_pages) * 0.5):
        return None  # index is mostly unnamed -- routing would be guesswork

    catalog = "\n".join(f"{p} = {drawing_index.get(str(p), f'Page {p}')}" for p in candidate_pages)
    sys_p = "You route construction questions to the correct drawing sheets. You output only JSON."
    usr_p = f"""A field question needs answering from a drawing set. Below is the sheet index.

Pick the sheets most likely to contain the answer. Think about which discipline and sheet type would carry this information (plans, schedules, details, RCPs, panel schedules, etc.). Include schedule sheets when the question involves equipment, finishes, doors, or power.

Return ONLY a JSON array of sheet page numbers, most relevant first, maximum {max_sheets} entries. Example: [12, 45, 46]

QUESTION: {question}

SHEET INDEX (page number = sheet name):
{catalog}"""
    try:
        res = ai_client.models.generate_content(
            model=MODEL_NAME,
            contents=[usr_p],
            config=types.GenerateContentConfig(
                system_instruction=sys_p,
                temperature=0.1,
                response_mime_type="application/json"
            )
        )
        picked = json.loads(res.text.strip().replace('```json', '').replace('```', '').strip())
        valid = [int(p) for p in picked if int(p) in candidate_pages]
        return valid[:max_sheets] or None
    except Exception:
        return None  # routing is an optimization -- fall back to full scan

def ask_drawings(file_bytes, pages, question, drawing_index, prior_turns=None, batch_size=8):
    """Answers a question against the given sheets, batching to stay memory-safe."""
    batches = [pages[i:i + batch_size] for i in range(0, len(pages), batch_size)]
    total_batches = len(batches)
    findings = []

    context = ""
    if prior_turns:
        recent = prior_turns[-3:]
        context = "\n\nEARLIER IN THIS CONVERSATION (for context on follow-up questions):\n" + "\n".join(
            f"Q: {t['q']}\nA: {t['a'][:400]}" for t in recent
        )

    progress_bar = st.progress(0)
    status_text = st.empty()

    for b_idx, batch in enumerate(batches):
        labels = ", ".join(drawing_index.get(str(p), f"Page {p}") for p in batch)
        status_text.markdown(f"**Reading sheets {b_idx + 1}/{total_batches}: {labels[:90]}...**")

        usr_p = (
            f"QUESTION FROM THE FIELD: {question}{context}\n\n"
            f"The attached images are these sheets, in order: {labels}.\n"
            f"Refer to sheets by these names when citing.\n\n"
            f"Answer the question from these sheets. If the answer is not on these particular "
            f"sheets, respond with exactly: NOT_ON_THESE_SHEETS"
        )
        imgs = [usr_p]
        for p_num in batch:
            imgs.append(convert_single_page(file_bytes, p_num))

        try:
            res = ai_client.models.generate_content(
                model=MODEL_NAME,
                contents=imgs,
                config=types.GenerateContentConfig(system_instruction=QA_SYSTEM_PROMPT, temperature=0.1)
            )
            txt = res.text.strip()
            if "NOT_ON_THESE_SHEETS" not in txt.upper():
                findings.append(txt)
        except Exception as e:
            status_text.warning(f"Sheet group {b_idx + 1} couldn't be read: {e}")

        del imgs
        gc.collect()
        progress_bar.progress(int(((b_idx + 1) / total_batches) * 90))

    progress_bar.progress(100)
    status_text.empty()

    if not findings:
        return ("Not found on the sheets searched. Try selecting additional sheets, or run the "
                "AI Drawing Indexer first so the assistant can route your question to the right sheets.")
    if len(findings) == 1:
        return findings[0]

    # Several sheet groups had partial answers -- reconcile into one
    merged = "\n\n---\n\n".join(findings)
    try:
        res = ai_client.models.generate_content(
            model=MODEL_NAME,
            contents=[f"QUESTION: {question}\n\nPartial answers found on different sheets:\n\n{merged}\n\n"
                      f"Combine these into one direct answer. Keep every sheet citation. If the partial "
                      f"answers conflict, say so explicitly and flag it as a coordination issue."],
            config=types.GenerateContentConfig(system_instruction=QA_SYSTEM_PROMPT, temperature=0.1)
        )
        return res.text
    except Exception:
        return merged

def create_pdf_report(project_name, content, title):
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("helvetica", 'B', 16)
    pdf.cell(190, 10, f"Project: {project_name}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
    pdf.set_font("helvetica", 'I', 12)
    pdf.cell(190, 10, f"Report: {title} | Generated: {datetime.now().strftime('%Y-%m-%d')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
    pdf.line(10, 30, 200, 30)
    pdf.ln(10)
    
    def safe_write(text):
        pdf.set_font("helvetica", '', 10)
        clean_t = str(text).replace('**', '').replace('### ', '').replace('## ', '').replace('# ', '')
        encoded_t = clean_t.encode('latin-1', 'ignore').decode('latin-1').replace('\t', '    ')
        words = [w[:80] + '-' + w[80:] if len(w) > 80 else w for w in encoded_t.split(' ')]
        pdf.multi_cell(190, 6, text=' '.join(words))

    if isinstance(content, list): safe_write("\n".join([f"- {item}" for item in content]))
    else: safe_write(str(content))
        
    return bytes(pdf.output())

# ==========================================
# BATCHED AI PROCESSING (memory-safe rewrite)
# ==========================================
# WHY THIS CHANGED:
# The original version converted every selected page to an image and held ALL of them
# in memory at once before making a single Gemini call. On a 100+ sheet set, that could
# push past Streamlit Cloud's free-tier memory ceiling (~1GB) and get the process killed
# by the OS with no Python traceback -- which is exactly the "blank crash" behavior we
# were debugging.
#
# This version processes pages in small batches. Each batch is converted, sent to Gemini,
# and then explicitly released from memory before the next batch starts. If the job needed
# more than one batch, a final lightweight TEXT-ONLY pass (no images, so it's cheap) merges
# and de-duplicates the batch results into one consolidated output in the original format.
#
# batch_size=15 is a starting point. If you still see crashes, lower it (e.g. 10). If it's
# fast and stable, you can raise it.
def run_ai_with_progress(file_bytes, target_pages, sys_prompt, usr_prompt, success_message="Task Complete!", batch_size=15):
    progress_bar = st.progress(0)
    status_text = st.empty()

    batches = [target_pages[i:i + batch_size] for i in range(0, len(target_pages), batch_size)]
    total_batches = len(batches)
    batch_outputs = []

    for b_idx, batch in enumerate(batches):
        status_text.markdown(f"**⚙️ Processing Batch {b_idx + 1}/{total_batches} (Sheets {batch[0]}-{batch[-1]})...**")

        d_imgs = [usr_prompt]
        for p_num in batch:
            d_imgs.append(convert_single_page(file_bytes, p_num))

        # Tell the model it's only seeing a slice of the full set, so it doesn't assume
        # completeness or reference sheets outside this batch.
        if total_batches > 1:
            d_imgs[0] = (
                usr_prompt
                + f"\n\n[NOTE: This is batch {b_idx + 1} of {total_batches}, covering sheets "
                  f"{batch[0]} through {batch[-1]} only. Extract findings ONLY from these sheets. "
                  f"The rest of the set is being processed in separate batches and merged afterward.]"
            )

        try:
            response = ai_client.models.generate_content(
                model=MODEL_NAME,
                contents=d_imgs,
                config=types.GenerateContentConfig(
                    system_instruction=sys_prompt,
                    temperature=0.2
                )
            )
            batch_outputs.append(response.text)
        except Exception as e:
            status_text.warning(f"⚠️ Batch {b_idx + 1} (sheets {batch[0]}-{batch[-1]}) failed: {e}")
            batch_outputs.append(f"[Batch {b_idx + 1}, sheets {batch[0]}-{batch[-1]} failed to process: {e}]")

        # Explicitly release this batch's images before moving on
        del d_imgs
        gc.collect()

        progress_bar.progress(int(((b_idx + 1) / total_batches) * 85))

    # Only one batch was needed -- nothing to merge, return as-is
    if total_batches == 1:
        progress_bar.progress(100)
        status_text.success(f"✅ {success_message}")
        return batch_outputs[0]

    # Multiple batches -- consolidate into one final result (text-only pass, no images)
    status_text.markdown("**🧠 Consolidating results across all batches...**")
    combined = "\n\n---BATCH BREAK---\n\n".join(
        f"[Batch {i + 1}, sheets {b[0]}-{b[-1]}]\n{out}"
        for i, (b, out) in enumerate(zip(batches, batch_outputs))
    )
    reduce_sys_prompt = (
        sys_prompt
        + "\n\nYou are now merging raw findings that were already extracted from separate "
          "batches of the same drawing set. Combine them into ONE consolidated result in the "
          "exact same output format as the original instructions. Remove duplicate or "
          "near-duplicate entries that appear across batches, but do not drop unique findings."
    )
    reduce_usr_prompt = f"Merge and de-duplicate these batch findings into a single final result:\n\n{combined}"

    try:
        final_response = ai_client.models.generate_content(
            model=MODEL_NAME,
            contents=[reduce_usr_prompt],
            config=types.GenerateContentConfig(system_instruction=reduce_sys_prompt, temperature=0.1)
        )
        result_text = final_response.text
    except Exception as e:
        status_text.warning(f"⚠️ Consolidation pass failed: {e}. Showing unmerged batch results instead.")
        result_text = combined

    progress_bar.progress(100)
    status_text.success(f"✅ {success_message}")
    return result_text

# ==========================================
# 5. SESSION INITIALIZATION
# ==========================================
keys_to_initialize = [
    'audit_results', 'takeoff_results', 'schedule_results', 'schedule_csv', 
    'doc_intel_results', 'est_results', 'submittal_results', 'drawing_index', 
    'audit_history', 'takeoff_history', 'schedule_history', 'intel_history', 
    'est_history', 'submittal_history', 'current_file', 'loaded_save_id',
    'qa_history'
]

for key in keys_to_initialize:
    if key not in st.session_state: 
        if 'history' in key or key in ['audit_results', 'takeoff_results', 'submittal_results']:
            st.session_state[key] = []
        elif key == 'drawing_index':
            st.session_state[key] = {}
        else:
            st.session_state[key] = ""

# ==========================================
# 6. SIDEBAR & SAVE SYSTEM
# ==========================================
with st.sidebar:
    st.markdown("## Site Pilot")
    st.caption("FIELD OS — V24.0")
    st.divider()
    
    st.markdown("### Document Uploads")
    uploaded_file = st.file_uploader("1️⃣ Base Drawings (.pdf)", type=["pdf"])
    spec_file = st.file_uploader("2️⃣ Project Specs (.pdf)", type=["pdf"])
    doc_file = st.file_uploader("3️⃣ Legal/Contracts (.pdf)", type=["pdf"])
    
    st.divider()
    st.markdown("### Save & Restore")
    save_file = st.file_uploader("4️⃣ Restore Project (.json)", type=["json"], help="Upload a previously downloaded save file here to restore your work.")
    
    # schedule_csv now holds binary .xlsx bytes (not text), which isn't
    # JSON-serializable -- excluded from the save file. After restoring a
    # project, just click "Generate Gantt Chart" again to rebuild it from
    # the preserved schedule_results.
    export_state = {k: st.session_state[k] for k in keys_to_initialize if k in st.session_state and k not in ('loaded_save_id', 'schedule_csv')}
    json_state = json.dumps(export_state)
    
    st.download_button(
        label="Download Save File",
        data=json_state,
        file_name=f"SitePilot_Save_{datetime.now().strftime('%Y%m%d')}.json",
        mime="application/json",
        type="primary"
    )

# ==========================================
# 7. CLOUD MEMORY LOGIC
# ==========================================
if uploaded_file and st.session_state.current_file != uploaded_file.name:
    st.session_state.current_file = uploaded_file.name
    if save_file is None or st.session_state.loaded_save_id != save_file.file_id:
        st.session_state.drawing_index = {}
        for h in ['audit', 'takeoff', 'schedule', 'est', 'intel', 'submittal']:
            st.session_state[f"{h}_history"] = []
            st.session_state[f"{h}_results"] = [] if h in ['audit', 'takeoff', 'submittal'] else ""
        st.session_state.schedule_csv = ""
        st.session_state.qa_history = []
    st.rerun()

if save_file and st.session_state.loaded_save_id != save_file.file_id:
    try:
        saved_data = json.loads(save_file.getvalue().decode("utf-8"))
        for k, v in saved_data.items():
            st.session_state[k] = v
        st.session_state.loaded_save_id = save_file.file_id
        st.success("✅ Project state restored successfully!")
        st.rerun()
    except Exception as e:
        st.error("🚨 Invalid save file.")

# ==========================================
# 8. MAIN LOGIC
# ==========================================
if uploaded_file:
    file_bytes = uploaded_file.read()
    total_pages = get_pdf_info(file_bytes)
    
    if not st.session_state.drawing_index:
        st.session_state.drawing_index = {str(i): f"Page {i}" for i in range(1, total_pages + 1)}

    st.markdown(f'''
        <div class="title-block">
            <div class="tb-field tb-brand"><div class="tb-value">SP</div></div>
            <div class="tb-field"><div class="tb-label">Project</div><div class="tb-value">{st.session_state.current_file}</div></div>
            <div class="tb-field"><div class="tb-label">Sheet Count</div><div class="tb-value">{total_pages}</div></div>
            <div class="tb-field"><div class="tb-label">Date</div><div class="tb-value">{datetime.now().strftime('%m.%d.%Y')}</div></div>
        </div>
    ''', unsafe_allow_html=True)

    with st.expander("AI Drawing Indexer", expanded=False):
        st.markdown("Extract sheet names from title blocks to automatically rename dropdown menus.")
        c_idx1, c_idx2 = st.columns([1, 4])
        with c_idx1:
            if st.button("Run Auto-Index"):
                idx_prog = st.progress(0); idx_stat = st.empty(); new_index = {}
                for i in range(1, total_pages + 1):
                    img = convert_single_page(file_bytes, i)
                    try:
                        usr_prompt = "Extract the Sheet Number and Sheet Title from this title block. Output ONLY in this exact format: 'SheetNumber - SheetTitle'."
                        sys_prompt = "You are a meticulous document archivist. Output strictly the requested format with no other text."
                        res = ai_client.models.generate_content(
                            model=MODEL_NAME, 
                            contents=[usr_prompt, img],
                            config=types.GenerateContentConfig(system_instruction=sys_prompt, temperature=0.1)
                        )
                        new_index[str(i)] = res.text.strip().replace('\n', '')
                    except: new_index[str(i)] = f"Page {i}"
                    idx_prog.progress(int((i / total_pages) * 100))
                st.session_state.drawing_index = new_index
                idx_stat.success("✅ Indexing Complete!")
                st.rerun()

    page_opts = list(st.session_state.drawing_index.values())
    tab_vdc, tab_est, tab_admin, tab_qa = st.tabs(["01 · Plan Room & VDC", "02 · Estimating & Docs", "03 · Admin & Specs", "04 · Ask the Drawings"])

    # --- TAB 1: PLAN ROOM ---
    with tab_vdc:
        st.markdown('<div class="tool-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Workspace Setup</div>', unsafe_allow_html=True)
        c_sel1, c_sel2 = st.columns([3, 1])
        with c_sel1:
            all_selected = st.checkbox("☑️ Select Entire Drawing Set")
            default_selection = page_opts if all_selected else []
            target_docs = st.multiselect("Target Sheets:", page_opts, default=default_selection, label_visibility="collapsed")
        with c_sel2: 
            st.markdown('<div class="btn-clear">', unsafe_allow_html=True)
            if st.button("Clear Workspace"):
                st.session_state.audit_results = []; st.session_state.takeoff_results = []
                st.session_state.schedule_results = ""; st.session_state.schedule_csv = ""
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        c_view, c_tools = st.columns([1.5, 1])
        with c_view:
            st.markdown("#### Sheet Viewer")
            selected_main = st.selectbox("Active View:", page_opts, label_visibility="collapsed")
            main_idx = int([k for k, v in st.session_state.drawing_index.items() if v == selected_main][0])
            st.markdown(f'<div class="ref-header">{selected_main}</div>', unsafe_allow_html=True)
            st.image(convert_single_page(file_bytes, main_idx), width='stretch')

        with c_tools:
            st.markdown("#### VDC Engines")
            # Clash Engine
            st.markdown('<div class="tool-card" style="padding: 15px;">', unsafe_allow_html=True)
            st.markdown('<div class="module-tag">Module 01 — Clash Audit</div>', unsafe_allow_html=True)
            if st.button("Run Clash Audit"):
                if target_docs:
                    p_scan = [int([k for k, v in st.session_state.drawing_index.items() if v == d][0]) for d in target_docs]
                    
                    sys_prompt = """You are a Master MEP Coordinator and Veteran Commercial Superintendent with 25 years of field experience. Your job is to audit commercial construction drawings and identify massive, expensive, project-halting constructability issues before they hit the field.
                    CRITICAL DIRECTIVE: You must strictly IGNORE minor drafting errors, text overlaps, spelling mistakes, or cosmetic issues. Do not waste time on fluff. 
                    FOCUS EXCLUSIVELY ON THESE 6 CRITICAL FAILURE POINTS:
                    1. Phasing & Scope Contradictions (The 'Frankenstein' Rule): Look for contradictions between "Existing/Demo" plans and "New Work" plans within the same trade. Flag if new approved drawings overlap, duplicate, or contradict existing approved drawings. 
                    2. Equipment vs. MEP Disconnects: Verify that heavy commercial equipment (e.g., kitchen hoods, RTUs, specialized machinery) has the correct and corresponding electrical (voltage/phase), plumbing, gas, and structural support on the MEP sheets. Flag missing utility connections.
                    3. Spatial Clashes & Interferences: Hunt for physical collisions. Look for ductwork, grease routing, or plumbing trenches intersecting footings, steel beams, shear walls, or load-bearing elements. Check if drop ceilings leave enough plenum space for specified HVAC equipment.
                    4. Utility Capacity & Load Deficiencies: Flag potential overloading of existing utility infrastructure. Look for new heavy equipment being added to existing electrical panels without load calculations, or undersized water/gas lines for the specified fixture counts.
                    5. Clearance, Code, & Life Safety: Hunt for missing working clearances around electrical panels and mechanical equipment. Flag ADA clearance violations, egress paths blocked by door swings, or missing fire-rated partition details.
                    6. Missing Critical Dimensions & Details: Identify areas where the field team cannot physically build because important measurements are missing.
                    OUTPUT FORMAT: Output only the major, expensive, schedule-killing issues. Start every single finding strictly with 'ISSUE: '. Be brutal, brief, and highly specific to the sheets and equipment tags provided."""
                    
                    usr_prompt = "Cross-reference the attached drawing sheets. Apply your 'Frankenstein Rule' and the other 5 critical failure points to hunt for scope, phasing, and physical contradictions. List the critical issues."
                    
                    res = run_ai_with_progress(file_bytes, p_scan, sys_prompt, usr_prompt, "Audit Complete!")
                    st.session_state.audit_results = [l.replace("ISSUE:", "").strip() for l in res.split("\n") if "ISSUE:" in l]
                    st.session_state.audit_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": "Audit", "results": st.session_state.audit_results})
                else: st.warning("Please select sheets first.")
            if st.session_state.audit_results:
                st.markdown('<div class="report-box accent-safety" style="padding: 10px;">', unsafe_allow_html=True)
                for issue in st.session_state.audit_results: st.write(f"🚩 {issue}")
                st.download_button("📥 Export PDF", create_pdf_report(st.session_state.current_file, st.session_state.audit_results, "Clash Audit"), "Audit.pdf", "application/pdf")
                st.markdown('</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Takeoff Engine
            st.markdown('<div class="tool-card" style="padding: 15px;">', unsafe_allow_html=True)
            st.markdown('<div class="module-tag">Module 02 — Material Takeoff</div>', unsafe_allow_html=True)
            if st.button("Material Takeoff"):
                if target_docs:
                    p_scan = [int([k for k, v in st.session_state.drawing_index.items() if v == d][0]) for d in target_docs]
                    sys_prompt = "You are a Senior Quantity Surveyor. Perform a highly accurate, structured material takeoff from the provided drawings."
                    usr_prompt = "Perform detailed material takeoff. Output continuous lines starting with 'TAKEOFF: '."
                    res = run_ai_with_progress(file_bytes, p_scan, sys_prompt, usr_prompt, "Takeoff Complete!")
                    st.session_state.takeoff_results = [l.replace("TAKEOFF:", "").strip() for l in res.split("\n") if "TAKEOFF:" in l]
                    st.session_state.takeoff_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": "Takeoff", "results": st.session_state.takeoff_results})
                else: st.warning("Please select sheets first.")
            if st.session_state.takeoff_results:
                st.markdown('<div class="report-box accent-verified" style="padding: 10px;">', unsafe_allow_html=True)
                for item in st.session_state.takeoff_results: st.write(f"📦 {item}")
                st.download_button("📥 Export PDF", create_pdf_report(st.session_state.current_file, st.session_state.takeoff_results, "Takeoff"), "Takeoff.pdf", "application/pdf")
                st.markdown('</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Timeline Engine
            st.markdown('<div class="tool-card" style="padding: 15px;">', unsafe_allow_html=True)
            st.markdown('<div class="module-tag">Module 03 — Project Timeline</div>', unsafe_allow_html=True)
            schedule_start_date = st.date_input("Project Start Date", value=datetime.now().date(), key="schedule_start_date")
            if st.button("Project Timeline"):
                if target_docs:
                    p_scan = [int([k for k, v in st.session_state.drawing_index.items() if v == d][0]) for d in target_docs]
                    sys_prompt = "You are a Master Project Scheduler specializing in commercial construction logic."
                    usr_prompt = f"Analyze drawings. The project start date is {schedule_start_date.strftime('%b %d, %Y')}. Generate a projected chronological construction timeline starting from this date, broken into discrete sequential tasks."
                    st.session_state.schedule_results = run_ai_with_progress(file_bytes, p_scan, sys_prompt, usr_prompt, "Timeline Generated!")
                    st.session_state.schedule_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": "Timeline", "results": st.session_state.schedule_results})
                else: st.warning("Please select sheets first.")
            if st.session_state.schedule_results:
                st.markdown('<div class="report-box" style="padding: 10px;">', unsafe_allow_html=True)
                st.markdown(st.session_state.schedule_results)
                if st.button("Generate Gantt Chart (.xlsx)", key="gen_gantt"):
                    with st.spinner("Building Gantt chart from your template..."):
                        try:
                            tasks = extract_tasks_json(st.session_state.schedule_results)
                            xlsx_bytes, truncated = fill_gantt_template(schedule_start_date, tasks)
                            st.session_state.schedule_csv = xlsx_bytes
                            if truncated:
                                st.warning(f"This timeline has more than {GANTT_MAX_TASK_ROWS} tasks — the template supports {GANTT_MAX_TASK_ROWS} rows, so it was truncated to fit. Consider grouping smaller tasks together.")
                        except Exception as e:
                            st.error(f"Couldn't build the Gantt chart: {e}")
                        st.rerun()
                if st.session_state.schedule_csv:
                    st.download_button("Download Gantt Chart (.xlsx)", st.session_state.schedule_csv, "Project_Schedule.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.markdown('</div>', unsafe_allow_html=True)

        st.divider()
        with st.expander("🗄️ VDC Archives (Recall Past Scans)"):
            ha1, ha2, ha3 = st.columns(3)
            with ha1:
                st.markdown("**Clash Audits**")
                for i, e in enumerate(st.session_state.audit_history):
                    with st.popover(f"🕒 {e['time']} Audit"):
                        for item in e['results']: st.write(f"- {item}")
                        st.download_button("📥 Download PDF", create_pdf_report(st.session_state.current_file, e['results'], "Audit"), f"Audit_{i}.pdf", key=f"dla_{i}")
            with ha2:
                st.markdown("**Material Takeoffs**")
                for i, e in enumerate(st.session_state.takeoff_history):
                    with st.popover(f"🕒 {e['time']} Takeoff"):
                        for item in e['results']: st.write(f"- {item}")
                        st.download_button("📥 Download PDF", create_pdf_report(st.session_state.current_file, e['results'], "Takeoff"), f"Takeoff_{i}.pdf", key=f"dlt_{i}")
            with ha3:
                st.markdown("**Project Timelines**")
                for i, e in enumerate(st.session_state.schedule_history):
                    with st.popover(f"🕒 {e['time']} Timeline"):
                        st.markdown(e['results'])
                        st.download_button("📥 Download PDF", create_pdf_report(st.session_state.current_file, e['results'], "Timeline"), f"Timeline_{i}.pdf", key=f"dls_{i}")

    # --- TAB 2: ESTIMATING ---
    with tab_est:
        col_est, col_doc = st.columns([1.2, 1])
        with col_est:
            st.markdown('<div class="tool-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">AI Estimator</div>', unsafe_allow_html=True)
            loc_multiplier = st.selectbox("Pricing Region:", ["National Average", "DMV Area (DC/MD/VA)", "New York", "Southeast"])
            st.markdown('<div class="module-tag">Module 04 — Baseline Estimate</div>', unsafe_allow_html=True)
            if st.button("Generate Baseline Estimate"):
                if target_docs:
                    p_scan = [int([k for k, v in st.session_state.drawing_index.items() if v == d][0]) for d in target_docs]
                    sys_prompt = f"You are a Chief Estimator for a massive commercial GC. Your pricing region multiplier logic is based on: {loc_multiplier}."
                    usr_prompt = "Generate a trade-grouped baseline estimate with line items and a budget summary. Format: Markdown."
                    st.session_state.est_results = run_ai_with_progress(file_bytes, p_scan, sys_prompt, usr_prompt, "Estimate Complete!")
                    st.session_state.est_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": loc_multiplier, "results": st.session_state.est_results})
                else: st.warning("Please return to the VDC tab and select target sheets.")
            if st.session_state.est_results:
                st.markdown(f'<div class="report-box">{st.session_state.est_results}</div>', unsafe_allow_html=True)
                st.download_button("📥 Export Estimate PDF", create_pdf_report(st.session_state.current_file, st.session_state.est_results, "Estimate"), "Estimate.pdf", "application/pdf")
            st.markdown('</div>', unsafe_allow_html=True)

        with col_doc:
            st.markdown('<div class="tool-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">Document Intelligence</div>', unsafe_allow_html=True)
            if doc_file:
                st.markdown('<div class="module-tag">Module 05 — Document Intelligence</div>', unsafe_allow_html=True)
                if st.button("Analyze Document"):
                    d_bytes = doc_file.read()
                    p_scan = list(range(1, get_pdf_info(d_bytes) + 1))
                    sys_prompt = "You are a Senior Construction Attorney and Risk Manager."
                    usr_prompt = "Summarize the primary purpose, key data points, financial impacts, and critical risks in this document."
                    st.session_state.doc_intel_results = run_ai_with_progress(d_bytes, p_scan, sys_prompt, usr_prompt, "Document Scanned!")
                    st.session_state.intel_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": doc_file.name, "results": st.session_state.doc_intel_results})
                if st.session_state.doc_intel_results: 
                    st.markdown(f'<div class="report-box">{st.session_state.doc_intel_results}</div>', unsafe_allow_html=True)
                    st.download_button("📥 Export Summary PDF", create_pdf_report(st.session_state.current_file, st.session_state.doc_intel_results, "Doc Summary"), "Summary.pdf", "application/pdf")
            else: st.info("Upload a secondary PDF to Slot 3 in the sidebar.")
            st.markdown('</div>', unsafe_allow_html=True)

        st.divider()
        with st.expander("🗄️ Estimating & Doc Archives"):
            ea1, ea2 = st.columns(2)
            with ea1:
                st.markdown("**Estimates**")
                for i, e in enumerate(st.session_state.est_history):
                    with st.popover(f"🕒 {e['time']} | {e['desc']}"):
                        st.markdown(e['results'])
                        st.download_button("📥 PDF", create_pdf_report(st.session_state.current_file, e['results'], "Estimate"), f"Est_{i}.pdf", key=f"dle_{i}")
            with ea2:
                st.markdown("**Document Summaries**")
                for i, e in enumerate(st.session_state.intel_history):
                    with st.popover(f"🕒 {e['time']} | {e['desc']}"):
                        st.markdown(e['results'])
                        st.download_button("📥 PDF", create_pdf_report(st.session_state.current_file, e['results'], "Summary"), f"Doc_{i}.pdf", key=f"dli_{i}")

    # --- TAB 3: ADMIN ---
    with tab_admin:
        st.markdown('<div class="tool-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Submittal Engine</div>', unsafe_allow_html=True)
        if spec_file:
            st.markdown('<div class="module-tag">Module 06 — Submittal Register</div>', unsafe_allow_html=True)
            if st.button("Generate Submittal Register"):
                s_bytes = spec_file.read(); s_total = get_pdf_info(s_bytes)
                p_scan = list(range(1, s_total + 1, 10))
                sys_prompt = "You are a Senior Project Engineer. Your job is to strictly extract submittal requirements from the project specifications."
                usr_prompt = "List required Shop Drawings, Product Data, and Samples. Start each with 'SUBMITTAL: '."
                res = run_ai_with_progress(s_bytes, p_scan, sys_prompt, usr_prompt, "Register Generated!")
                st.session_state.submittal_results = [l.replace("SUBMITTAL:", "").strip() for l in res.split("\n") if "SUBMITTAL:" in l]
                st.session_state.submittal_history.insert(0, {"time": datetime.now().strftime("%I:%M %p"), "desc": "Scan", "results": st.session_state.submittal_results})
            if st.session_state.submittal_results:
                st.markdown('<div class="report-box accent-safety">', unsafe_allow_html=True)
                for s in st.session_state.submittal_results: st.write(f"📁 {s}")
                st.markdown('</div>', unsafe_allow_html=True)
                st.download_button("📥 Export Submittal Log PDF", create_pdf_report(st.session_state.current_file, st.session_state.submittal_results, "Submittal Register"), "Submittals.pdf", "application/pdf")
        else: st.info("Upload Specifications in Slot 2 to enable Submittal scanning.")
        st.markdown('</div>', unsafe_allow_html=True)
        
        st.divider()
        with st.expander("🗄️ Admin Archives"):
            st.markdown("**Submittal Registers**")
            for i, e in enumerate(st.session_state.submittal_history):
                with st.popover(f"🕒 {e['time']} Register"):
                    for item in e['results']: st.write(f"- {item}")
                    st.download_button("📥 PDF", create_pdf_report(st.session_state.current_file, e['results'], "Submittal Log"), f"Sub_{i}.pdf", key=f"dls_{i}")

    # --- TAB 4: ASK THE DRAWINGS ---
    with tab_qa:
        st.markdown('<div class="tool-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Ask the Drawings</div>', unsafe_allow_html=True)
        st.markdown('<div class="module-tag">Module 07 — Field Q&amp;A</div>', unsafe_allow_html=True)
        st.caption("Ask about dimensions, areas, finishes, equipment, power requirements, clearances — "
                   "anything on the sheets. Answers cite the sheet they came from.")

        indexed = any(not str(v).strip().lower().startswith("page ") for v in st.session_state.drawing_index.values())

        c_qa1, c_qa2 = st.columns([3, 1])
        with c_qa1:
            smart_route = st.checkbox(
                "Smart sheet routing (recommended)",
                value=indexed,
                disabled=not indexed,
                help="Finds the sheets most likely to hold the answer before reading them. "
                     "Requires running the AI Drawing Indexer first so sheets have real names."
            )
        with c_qa2:
            st.markdown('<div class="btn-clear">', unsafe_allow_html=True)
            if st.button("Clear Conversation", key="clear_qa"):
                st.session_state.qa_history = []
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        if not indexed:
            st.info("Run the **AI Drawing Indexer** above to enable smart routing — without it, "
                    "every question scans all your selected sheets, which is slower on a large set.")

        with st.form("qa_form", clear_on_submit=True):
            question = st.text_input(
                "Question",
                placeholder="e.g. What are the dimensions of the walk-in cooler?",
                label_visibility="collapsed"
            )
            asked = st.form_submit_button("Ask")

        if asked and question.strip():
            if not target_docs:
                st.warning("Select target sheets in the Plan Room tab first (or check 'Select Entire Drawing Set').")
            else:
                candidate_pages = [int([k for k, v in st.session_state.drawing_index.items() if v == d][0]) for d in target_docs]
                search_pages = candidate_pages
                routed = None
                if smart_route and indexed:
                    with st.spinner("Locating relevant sheets..."):
                        routed = select_relevant_sheets(question, st.session_state.drawing_index, candidate_pages)
                    if routed:
                        search_pages = routed
                        names = ", ".join(st.session_state.drawing_index.get(str(p), f"Page {p}") for p in routed)
                        st.caption(f"Searching: {names}")

                answer = ask_drawings(
                    file_bytes, search_pages, question.strip(),
                    st.session_state.drawing_index,
                    prior_turns=st.session_state.qa_history
                )
                st.session_state.qa_history.append({
                    "time": datetime.now().strftime("%I:%M %p"),
                    "q": question.strip(),
                    "a": answer,
                    "sheets": ", ".join(st.session_state.drawing_index.get(str(p), f"Page {p}") for p in search_pages)
                })
                st.rerun()

        # Conversation, most recent first
        for turn in reversed(st.session_state.qa_history):
            st.markdown(
                f'<div class="ref-header" style="border-radius:3px; margin-top:16px;">{turn["q"]}</div>',
                unsafe_allow_html=True
            )
            st.markdown('<div class="report-box" style="margin-top:0;">', unsafe_allow_html=True)
            st.markdown(turn["a"])
            st.caption(f"{turn['time']} · sheets searched: {turn['sheets']}")
            st.markdown('</div>', unsafe_allow_html=True)

        if st.session_state.qa_history:
            transcript = "\n\n".join(
                f"Q: {t['q']}\nSheets: {t['sheets']}\n\nA: {t['a']}" for t in st.session_state.qa_history
            )
            st.download_button(
                "📥 Export Q&A Transcript",
                create_pdf_report(st.session_state.current_file, transcript, "Drawing Q&A"),
                "Drawing_QA.pdf", "application/pdf", key="dl_qa"
            )
        st.markdown('</div>', unsafe_allow_html=True)

# ==========================================
# 8. LANDING PAGE
# ==========================================
else:
    st.markdown('''
        <div style="text-align:center; padding:110px 20px 100px;">
            <div class="hero-eyebrow">Field Intelligence System</div>
            <h1 class="hero-title">Site Pilot AI</h1>
            <p class="hero-sub">Upload base drawings in the sidebar to initialize the project environment.</p>
        </div>
    ''', unsafe_allow_html=True)







