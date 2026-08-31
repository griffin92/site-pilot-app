"""Timeline generation and Gantt template fill.

DATE MATH LIVES IN PYTHON, NOT THE PROMPT. Language models are unreliable at
business-day arithmetic -- they'll cheerfully put a task start on a Saturday
or drift a chain of dependencies by days. The AI proposes task names,
durations, and dependencies; Python computes every actual date.
"""
import io
from datetime import timedelta

import openpyxl

from services import ai
from engines import prompts
from config import (GANTT_TEMPLATE_PATH, GANTT_TASK_START_ROW,
                    GANTT_MAX_TASK_ROWS, DEEP_SCAN_BATCH)


# ------------------------------------------------------------- date helpers

def add_business_days(start, work_days):
    """Mirrors the template's WORKDAY(start, work_days-1)."""
    if work_days <= 1:
        return start
    d, remaining = start, work_days - 1
    while remaining > 0:
        d += timedelta(days=1)
        if d.weekday() < 5:
            remaining -= 1
    return d


def next_business_day(d):
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def schedule_tasks(project_start, tasks):
    end_by_wbs = {}
    for t in tasks:
        preds = t.get("predecessors") or []
        pred_ends = [end_by_wbs[p] for p in preds if p in end_by_wbs]
        start = next_business_day(max(pred_ends) + timedelta(days=1)) if pred_ends else project_start
        end = add_business_days(start, max(1, int(t.get("work_days", 1))))
        t["start"], t["end"] = start, end
        end_by_wbs[t["wbs"]] = end
    return tasks


# ---------------------------------------------------------------- generation

def generate_timeline(file_bytes, pages, start_date, label_fn=None):
    user = (f"Analyze these drawings. Project start date is {start_date.strftime('%b %d, %Y')}. "
            f"Produce a chronological construction timeline broken into discrete sequential "
            f"tasks, grouped by phase, with a working-day duration and prerequisites for each.")
    return ai.batched_scan(
        file_bytes, pages, prompts.SCHEDULER, user,
        batch_size=DEEP_SCAN_BATCH, label_fn=label_fn,
        progress_label="Building timeline",
    )


def extract_tasks(timeline_text):
    """Free text -> structured task list via JSON mode."""
    user = f"""Extract every task from this construction timeline into a JSON array.
Each item must have exactly:
- "wbs": sequential integer starting at 1
- "task": short task name
- "work_days": duration in working days (integer, min 1)
- "predecessors": array of WBS integers it depends on (empty if none)

Output ONLY the raw JSON array.

Timeline:
{timeline_text}"""
    return ai.generate_json(
        [user],
        "You are a data extraction engine converting construction schedules into structured task lists.",
        default=[],
    )


def fill_gantt(project_start, tasks):
    """Writes only the plain input cells of the Vertex42 template; every
    locked formula (End, Cal Days, Days Done/Left) is left untouched so the
    workbook behaves exactly as the original does.

    Returns (xlsx_bytes, truncated_flag).
    """
    wb = openpyxl.load_workbook(GANTT_TEMPLATE_PATH)
    ws = wb["GanttChart"]
    ws["G6"] = project_start

    scheduled = schedule_tasks(project_start, tasks)
    truncated = len(scheduled) > GANTT_MAX_TASK_ROWS
    scheduled = scheduled[:GANTT_MAX_TASK_ROWS]

    for i in range(GANTT_MAX_TASK_ROWS):
        row = GANTT_TASK_START_ROW + i
        if i < len(scheduled):
            t = scheduled[i]
            ws.cell(row=row, column=1).value = t["wbs"]
            ws.cell(row=row, column=2).value = t["task"]
            ws.cell(row=row, column=4).value = ",".join(str(p) for p in t.get("predecessors", [])) or None
            ws.cell(row=row, column=7).value = t["start"]
            ws.cell(row=row, column=9).value = max(1, int(t.get("work_days", 1)))
            ws.cell(row=row, column=10).value = 0
        else:
            # openpyxl treats .cell(value=None) as a no-op, so clear via .value
            for col in (1, 2, 4, 7, 9, 10):
                ws.cell(row=row, column=col).value = None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), truncated
