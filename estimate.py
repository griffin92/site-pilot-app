"""Cost estimating.

DIVISION OF LABOUR -- this is the whole design:
  AI  -> identifies scope, quantities, units, and a ROM unit-cost STARTING POINT
  YOU -> edit any unit cost; the estimate becomes yours, not the model's
  PY  -> does every multiplication, subtotal, markup and total

Language models are unreliable at arithmetic and have no live pricing data.
Letting one produce a bottom-line number that looks authoritative is the most
dangerous thing this app could do, so it never computes money here. Every
figure below the unit-cost column is Python.
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services import ai
from engines import prompts
from config import DEEP_SCAN_BATCH

# Rough regional cost factors relative to US national average. These are
# planning-grade adjustments, not indexed data -- override them if you have
# better local numbers.
REGIONS = {
    "National Average": 1.00,
    "DMV (DC/MD/VA)": 1.11,
    "New York Metro": 1.32,
    "Boston": 1.22,
    "Chicago": 1.10,
    "Southeast": 0.92,
    "Texas": 0.94,
    "California (Bay Area)": 1.30,
    "Pacific Northwest": 1.08,
}

DEFAULT_MARKUPS = {
    "general_conditions": 8.0,
    "overhead": 6.0,
    "profit": 8.0,
    "contingency": 10.0,
}

LINE_SCHEMA = """Return ONLY a JSON array. Each element must have exactly these fields:
- "division": CSI division, e.g. "03 - Concrete", "09 - Finishes", "23 - HVAC" (string)
- "item": the scope item (string)
- "quantity": numeric quantity (a number, not a string)
- "unit": SF, LF, EA, CY, TON, LS, etc. (string)
- "unit_cost": ROM installed unit cost in US dollars, national average, material + labor (a number)
- "confidence": "High" if the quantity is dimensioned or scheduled on the drawings, "Medium" if derived from labeled dimensions, "Low" if the scope is implied but not quantified (string)
- "basis": one short sentence on where the quantity came from (string)
- "sheet": sheet name it came from (string)

Use national-average pricing; regional adjustment is applied separately.
Do not include markups, general conditions, overhead, profit, or contingency
as line items -- those are applied separately.
If nothing is quantifiable on these sheets, return an empty array: []"""


def generate_line_items(file_bytes, pages, label_fn=None):
    """Returns (line_items, warnings)."""
    user = ("Produce a trade-grouped ROM estimate basis from the attached sheets. "
            "Identify scope, quantities, and starting unit costs.\n\n" + LINE_SCHEMA)

    records, failed = ai.batched_json_scan(
        file_bytes, pages, prompts.ESTIMATOR, user,
        batch_size=DEEP_SCAN_BATCH, label_fn=label_fn,
        progress_label="Pricing scope",
    )

    clean = []
    for r in records:
        try:
            r["quantity"] = float(str(r.get("quantity", "")).replace(",", ""))
            r["unit_cost"] = float(str(r.get("unit_cost", "")).replace(",", "").replace("$", ""))
        except (TypeError, ValueError):
            continue
        if r["quantity"] <= 0:
            continue
        clean.append(r)

    clean.sort(key=lambda r: (str(r.get("division", "zz")), str(r.get("item", ""))))

    warnings = []
    if failed:
        warnings.append(f"{len(failed)} sheet group(s) returned unreadable output.")
    low = sum(1 for r in clean if str(r.get("confidence", "")).lower() == "low")
    if low:
        warnings.append(f"{low} of {len(clean)} line items are low-confidence — "
                        f"the scope is implied but not quantified on the drawings.")
    return clean, warnings


def from_takeoff(takeoff_items):
    """Seed an estimate from an existing takeoff so quantities aren't
    re-derived (and can't disagree between the two engines)."""
    out = []
    for t in takeoff_items:
        out.append({
            "division": t.get("division", "Unclassified"),
            "item": t.get("item", ""),
            "quantity": t.get("quantity", 0),
            "unit": t.get("unit", "EA"),
            "unit_cost": 0.0,          # user fills these in
            "confidence": "High" if str(t.get("basis", "")).upper() == "LABELED" else "Medium",
            "basis": f"From takeoff ({t.get('basis','')})",
            "sheet": t.get("sheet", ""),
        })
    return out


# --------------------------------------------------------------- the math
# All of this is deterministic Python. No model involvement.

def compute(line_items, region="National Average", markups=None, region_factor=None):
    """Returns a dict with per-line extended costs and the full rollup."""
    markups = {**DEFAULT_MARKUPS, **(markups or {})}
    factor = region_factor if region_factor is not None else REGIONS.get(region, 1.0)

    lines = []
    for r in line_items:
        try:
            qty = float(r.get("quantity", 0) or 0)
            rate = float(r.get("unit_cost", 0) or 0)
        except (TypeError, ValueError):
            qty, rate = 0.0, 0.0
        base = qty * rate
        lines.append({**r,
                      "quantity": qty,
                      "unit_cost": rate,
                      "extended": base,
                      "extended_regional": base * factor})

    direct = sum(l["extended_regional"] for l in lines)

    gc_amt = direct * markups["general_conditions"] / 100.0
    subtotal_gc = direct + gc_amt
    oh_amt = subtotal_gc * markups["overhead"] / 100.0
    subtotal_oh = subtotal_gc + oh_amt
    profit_amt = subtotal_oh * markups["profit"] / 100.0
    subtotal_profit = subtotal_oh + profit_amt
    cont_amt = subtotal_profit * markups["contingency"] / 100.0
    total = subtotal_profit + cont_amt

    by_div = {}
    for l in lines:
        d = str(l.get("division", "Unclassified"))
        by_div[d] = by_div.get(d, 0.0) + l["extended_regional"]

    return {
        "lines": lines,
        "region": region,
        "region_factor": factor,
        "markups": markups,
        "direct": direct,
        "gc_amt": gc_amt,
        "oh_amt": oh_amt,
        "profit_amt": profit_amt,
        "contingency_amt": cont_amt,
        "total": total,
        "by_division": dict(sorted(by_div.items())),
    }


# --------------------------------------------------------------- xlsx export

_NAVY = "0E1B2E"
_STEEL = "52627A"
_LIGHT = "EEF1F4"


def to_xlsx(result, project):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"

    thin = Side(style="thin", color="C8CED6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{project.get('name', 'Project')} — ROM Cost Estimate"
    ws["A1"].font = Font(size=15, bold=True, color=_NAVY)
    ws["A2"] = (f"Project #{project.get('number', '—')}   |   "
                f"Region: {result['region']} (x{result['region_factor']:.2f})   |   "
                f"Generated {datetime.now().strftime('%B %d, %Y')}")
    ws["A2"].font = Font(size=9, italic=True, color=_STEEL)
    ws["A3"] = ("ROM planning estimate. Unit costs require validation against current "
                "subcontractor pricing before any commitment.")
    ws["A3"].font = Font(size=9, bold=True, color="C0392B")

    headers = ["Division", "Item", "Qty", "Unit", "Unit Cost",
               "Extended", "Regional", "Confidence", "Sheet"]
    hrow = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = hrow + 1
    current_div = None
    for l in result["lines"]:
        div = str(l.get("division", "Unclassified"))
        if div != current_div:
            current_div = div
            ws.cell(row=r, column=1, value=div).font = Font(bold=True, size=9, color=_NAVY)
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=_LIGHT)
            r += 1

        ws.cell(row=r, column=2, value=l.get("item", ""))
        ws.cell(row=r, column=3, value=round(l["quantity"], 2))
        ws.cell(row=r, column=4, value=l.get("unit", ""))
        ws.cell(row=r, column=5, value=round(l["unit_cost"], 2)).number_format = '"$"#,##0.00'
        # Live formulas so the sheet stays editable after export
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = '"$"#,##0'
        ws.cell(row=r, column=7,
                value=f"=F{r}*{result['region_factor']}").number_format = '"$"#,##0'
        ws.cell(row=r, column=8, value=l.get("confidence", ""))
        ws.cell(row=r, column=9, value=l.get("sheet", ""))
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
        r += 1

    first_data, last_data = hrow + 1, r - 1
    r += 1

    def total_row(label, formula, bold=False, money=True):
        nonlocal r
        ws.cell(row=r, column=5, value=label).font = Font(bold=bold, size=10)
        cell = ws.cell(row=r, column=7, value=formula)
        cell.font = Font(bold=bold, size=10)
        if money:
            cell.number_format = '"$"#,##0'
        r += 1
        return r - 1

    direct_row = total_row("Direct Cost", f"=SUM(G{first_data}:G{last_data})", bold=True)
    m = result["markups"]
    gc_row = total_row(f"General Conditions ({m['general_conditions']}%)",
                       f"=G{direct_row}*{m['general_conditions'] / 100}")
    oh_row = total_row(f"Overhead ({m['overhead']}%)",
                       f"=(G{direct_row}+G{gc_row})*{m['overhead'] / 100}")
    pr_row = total_row(f"Profit ({m['profit']}%)",
                       f"=(G{direct_row}+G{gc_row}+G{oh_row})*{m['profit'] / 100}")
    ct_row = total_row(f"Contingency ({m['contingency']}%)",
                       f"=(G{direct_row}+G{gc_row}+G{oh_row}+G{pr_row})*{m['contingency'] / 100}")

    ws.cell(row=r, column=5, value="TOTAL").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=_NAVY)
    tot = ws.cell(row=r, column=7,
                  value=f"=G{direct_row}+G{gc_row}+G{oh_row}+G{pr_row}+G{ct_row}")
    tot.font = Font(bold=True, size=12, color="FFFFFF")
    tot.fill = PatternFill("solid", fgColor=_NAVY)
    tot.number_format = '"$"#,##0'

    for col, w in zip("ABCDEFGHI", [20, 46, 10, 8, 13, 14, 14, 12, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hrow + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
